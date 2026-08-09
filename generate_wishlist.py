#!/usr/bin/env python3
"""
Endgame Analysis -> tiered DIM wishlist generator.

Usage:
    export BUNGIE_API_KEY=your-key-here
    python generate_wishlist.py path/to/spreadsheet.xlsx -o wishlist.txt

Requires: pip install requests openpyxl

Reads the current-format weapon sheets (Name / Barrel / Mag / Perk 1 / Perk 2 /
Origin Trait / Tier), pulls the live Bungie manifest, resolves every weapon name
to all obtainable item hashes (incl. Adept/Timelost/Harrowed and reissues),
keeps only hashes whose plug sets can actually roll the listed trait perks, and
emits:

  God Roll  - barrel + mag + trait combo (+ verified origin trait)
  Good Roll - trait combo only (+ verified origin trait)

Filter in DIM:  wishlistnotes:"god roll"   wishlistnotes:"good roll"
                wishlistnotes:"good roll" -wishlistnotes:"god roll"
"""

import argparse
import itertools
import json
import os
import re
import sys
import tempfile
from collections import Counter, defaultdict

import requests
from openpyxl import load_workbook

SHEETS = ['Autos', 'Bows', 'HCs', 'Pulses', 'Scouts', 'Sidearms', 'SMGs',
          'BGLs', 'Fusions', 'Glaives', 'Shotguns', 'Snipers',
          'Rocket Sidearms', 'Traces', 'HGLs', 'LFRs', 'LMGs', 'Rockets',
          'Swords', 'Other']
ADEPT_SUFFIXES = [' (adept)', ' (timelost)', ' (harrowed)']
SKIP_VALUES = {'none', 'n/a', '?'}


def norm(s: str) -> str:
    """Normalize a name for matching: straight quotes, collapsed spaces, casefold."""
    return re.sub(r'\s+', ' ',
                  s.replace('\u2019', "'").replace('\u2018', "'").strip()).casefold()


def opts(cell) -> list[str]:
    """Split a newline-separated spreadsheet cell into option names."""
    return [x.strip() for x in (str(cell) if cell else '').split('\n')
            if x.strip() and x.strip().casefold() not in SKIP_VALUES]


# ---------------------------------------------------------------- manifest ---

def fetch_manifest(api_key: str, cache_dir: str) -> tuple[dict, dict]:
    """Download (with local caching) item and plug-set definitions."""
    h = {'X-API-Key': api_key}
    meta = requests.get('https://www.bungie.net/Platform/Destiny2/Manifest/',
                        headers=h, timeout=60).json()
    if meta.get('ErrorStatus') != 'Success':
        sys.exit(f"Manifest request failed: {meta.get('ErrorStatus')} — check API key")
    paths = meta['Response']['jsonWorldComponentContentPaths']['en']
    out = []
    for comp in ('DestinyInventoryItemDefinition', 'DestinyPlugSetDefinition'):
        path = paths[comp]
        cache = os.path.join(cache_dir, os.path.basename(path))
        if not os.path.exists(cache):
            print(f'Downloading {comp} ...', file=sys.stderr)
            r = requests.get('https://www.bungie.net' + path, timeout=600)
            r.raise_for_status()
            with open(cache, 'wb') as f:
                f.write(r.content)
        with open(cache) as f:
            out.append(json.load(f))
    return out[0], out[1]


def build_indexes(items: dict, plugsets_raw: dict):
    """Slim the manifest into the lookups we need."""
    slim, weapons_by_name = {}, defaultdict(list)
    for h, d in items.items():
        name = d.get('displayProperties', {}).get('name', '')
        entry = {
            'name': name,
            'icon': d.get('displayProperties', {}).get('icon', ''),
            'itemTypeDisplayName': d.get('itemTypeDisplayName', ''),
            'sockets': d.get('sockets') if d.get('itemType') == 3 else None,
        }
        slim[int(h)] = entry
        if d.get('itemType') == 3 and name and entry['sockets']:
            weapons_by_name[norm(name)].append(int(h))
    plugsets = {int(h): [p['plugItemHash'] for p in d.get('reusablePlugItems', [])]
                for h, d in plugsets_raw.items()}
    return slim, weapons_by_name, plugsets


def _plug_name_index(slim: dict, plugsets: dict, whash: int) -> dict[str, list]:
    """All plugs available on a weapon hash: normalized name -> [(enhanced, hash)]."""
    plugs = set()
    for se in slim[whash]['sockets'].get('socketEntries', []):
        for key in ('randomizedPlugSetHash', 'reusablePlugSetHash'):
            ps = se.get(key)
            if ps and ps in plugsets:
                plugs.update(plugsets[ps])
        if se.get('singleInitialItemHash'):
            plugs.add(se['singleInitialItemHash'])
        for rp in se.get('reusablePlugItems', []):
            plugs.add(rp['plugItemHash'])
    idx = defaultdict(list)
    for ph in plugs:
        pd = slim.get(ph)
        if not pd or not pd['name']:
            continue
        enhanced = 'enhanced' in (pd['itemTypeDisplayName'] or '').casefold()
        idx[norm(pd['name'])].append((enhanced, ph))
    return idx


def plug_index(slim: dict, plugsets: dict, whash: int) -> dict[str, int]:
    """All plugs available on a weapon hash: normalized name -> canonical hash.

    Prefers the base (non-enhanced) hash when both exist; DIM's wishlist
    matching treats enhanced traits as equal to their base version.
    """
    idx = _plug_name_index(slim, plugsets, whash)
    return {n: sorted(lst)[0][1] for n, lst in idx.items()}


# ------------------------------------------------------------- spreadsheet ---

def read_spreadsheet(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    rows_out = []
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f'WARNING: sheet {sheet!r} not found — skipped', file=sys.stderr)
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c else '' for c in rows[1]]
        idx = {k: (header.index(k) if k in header else None)
               for k in ['Name', 'Barrel', 'Mag', 'Perk 1', 'Perk 2',
                         'Origin Trait', 'Tier', 'Frame', 'Source']}
        if idx['Name'] is None or idx['Barrel'] is None:
            print(f'WARNING: sheet {sheet!r} has unexpected layout — skipped',
                  file=sys.stderr)
            continue
        for r in rows[2:]:
            if not r or idx['Name'] >= len(r) or not r[idx['Name']]:
                continue
            w = {'sheet': sheet}
            for k, i in idx.items():
                v = r[i] if (i is not None and i < len(r)) else None
                w[k] = str(v).strip() if v is not None else ''
            rows_out.append(w)
    return rows_out


# -------------------------------------------------------------- generation ---

def generate(rows, slim, weapons_by_name, plugsets):
    """Emit a strict quality ladder. DIM applies the FIRST matching line per
    item (its perks get badges, its note displays and is what wishlistnotes:
    filters see), so sections are ordered best-tier-first and each item's
    single visible tag is the best tier it achieves:

      1. Perfect Roll  - every listed trait + barrel + mag
      2. Full Set      - every listed trait, any barrel/mag
      3. Combo a+b     - a desired traits in col 3, b in col 4 (deepest first)
      4. God Roll      - one desired trait per column + barrel + mag
      5. Good Roll     - one desired trait per column
    """
    stats = Counter()

    # ---- pass 1: resolve every weapon row against every candidate hash ----
    resolved = []  # (row, hash, rb, rm, r1, r2, otail, rec_hint)
    for w in rows:
        base = w['Name'].split('\n')[0].strip()
        p1s, p2s = opts(w['Perk 1']), opts(w['Perk 2'])
        if not p1s and not p2s:
            continue
        barrels, mags = opts(w['Barrel']), opts(w['Mag'])
        origins = opts(w['Origin Trait'])
        cands = []
        for key in [norm(base)] + [norm(base) + s for s in ADEPT_SUFFIXES]:
            cands.extend(weapons_by_name.get(key, []))
        if not cands:
            stats['unmatched_weapons'] += 1
            print(f'UNMATCHED: [{w["sheet"]}] {base}', file=sys.stderr)
            continue
        for wh in cands:
            canon = plug_index(slim, plugsets, wh)
            rb = [canon[norm(o)] for o in barrels if norm(o) in canon]
            rm = [canon[norm(o)] for o in mags if norm(o) in canon]
            r1 = [canon[norm(o)] for o in p1s if norm(o) in canon]
            r2 = [canon[norm(o)] for o in p2s if norm(o) in canon]
            if (p1s and not r1) or (p2s and not r2):
                continue
            ro = next((canon[norm(o)] for o in origins if norm(o) in canon), None)
            # the recommended roll on THIS hash: the first-listed option per
            # column that the hash can actually roll (relevant on legacy hashes
            # that can't roll the sheet's true top pick)
            rec_hint = ' + '.join(next(o for o in src if norm(o) in canon)
                                  for src, res in ((p1s, r1), (p2s, r2)) if res)
            resolved.append((w, wh, rb, rm, r1, r2, [ro] if ro else [], rec_hint))

    # ---- pass 2: emit ladder sections ----
    lines = [
        'title:Endgame Analysis Wishlist — quality ladder',
        'description:From the Endgame Analysis spreadsheet. DIM applies the FIRST '
        'matching entry per item, so entries are ordered best-first and each item '
        'shows exactly one tag = its best tier: Perfect Roll (all listed traits + '
        'barrel/mag) > Full Set (all listed traits) > Combo a+b (a/b desired traits '
        'per column, deepest first) > God Roll (one per column + barrel/mag) > '
        'Good Roll. Single lone perks are not flagged. Filter: wishlistnotes:"perfect roll", '
        '"full set", "combo 3+3" etc., "god roll", "good roll". '
        'Every note also shows the sheet’s recommended roll; entries that ARE it '
        'read “★ RECOMMENDED ROLL” — filter wishlistnotes:"recommended roll". '
        'Masterworks not matched.',
        '',
    ]

    def tag_of(w):
        return f"Tier {(w['Tier'] or '?').strip()} {w['sheet']}"

    def section(title):
        lines.extend(['', f'// ============ {title} ============', ''])

    def emit(wh, perks, label, w, is_rec, rec_hint):
        # every note carries the recommendation: either "this IS it" (a
        # filterable tag) or a hint naming the top-pick traits
        extra = ' | ★ RECOMMENDED ROLL' if is_rec else f' | ★ Rec: {rec_hint}'
        lines.append(f"dimwishlist:item={wh}&perks={','.join(map(str, perks))}"
                     f"#notes:{label} — {tag_of(w)}{extra}")
        stats[label.lower().replace(' ', '_').replace('+', 'p')] += 1

    # An entry "is" the recommended roll when its traits include the
    # first-resolvable pick in each column AND any barrel/mag it names are the
    # first-resolvable ones. Entries that omit barrel/mag entirely (Full Set,
    # Combo, Good) can still be the recommendation — trait-wise it's all there.

    # 1. Perfect Roll
    section('1. PERFECT ROLL (all listed traits + barrel/mag)')
    for w, wh, rb, rm, r1, r2, otail, hint in resolved:
        if r1 and r2 and len(r1) <= 3 and len(r2) <= 3 and (rb or rm):
            for bc in (rb or [None]):
                for mc in (rm or [None]):
                    perks = [x for x in [bc, mc] if x] + r1 + r2 + otail
                    rec = (not rb or bc == rb[0]) and (not rm or mc == rm[0])
                    emit(wh, perks, 'Perfect Roll', w, rec, hint)

    # 2. Full Set
    section('2. FULL SET (all listed traits, any barrel/mag)')
    for w, wh, rb, rm, r1, r2, otail, hint in resolved:
        if r1 and r2 and len(r1) <= 3 and len(r2) <= 3:
            emit(wh, r1 + r2 + otail, 'Full Set', w, True, hint)

    # 3. Combo depths, deepest first
    depths = sorted(((a, b) for a in (1, 2, 3) for b in (1, 2, 3)
                     if (a, b) != (1, 1)), key=lambda ab: (-(ab[0] + ab[1]), -max(ab)))
    for a, b in depths:
        section(f'3. COMBO {a}+{b}')
        for w, wh, rb, rm, r1, r2, otail, hint in resolved:
            if len(r1) < a or len(r2) < b:
                continue
            for c1 in itertools.combinations(r1, a):
                for c2 in itertools.combinations(r2, b):
                    rec = r1[0] in c1 and r2[0] in c2
                    emit(wh, list(c1) + list(c2) + otail, f'Combo {a}+{b}', w,
                         rec, hint)

    # 4. God Roll
    section('4. GOD ROLL (one trait per column + barrel/mag)')
    for w, wh, rb, rm, r1, r2, otail, hint in resolved:
        if not (rb or rm):
            continue
        for t1 in (r1 or [None]):
            for t2 in (r2 or [None]):
                tc = [x for x in [t1, t2] if x]
                if not tc:
                    continue
                for bc in (rb or [None]):
                    for mc in (rm or [None]):
                        perks = [x for x in [bc, mc] if x] + tc + otail
                        rec = ((not r1 or t1 == r1[0]) and (not r2 or t2 == r2[0])
                               and (not rb or bc == rb[0]) and (not rm or mc == rm[0]))
                        emit(wh, perks, 'God Roll', w, rec, hint)

    # 5. Good Roll (includes weapons with no barrel/mag columns)
    section('5. GOOD ROLL (one trait per column, any barrel/mag)')
    for w, wh, rb, rm, r1, r2, otail, hint in resolved:
        label = 'Good Roll' if (rb or rm) else 'God Roll + Good Roll'
        for t1 in (r1 or [None]):
            for t2 in (r2 or [None]):
                tc = [x for x in [t1, t2] if x]
                if tc:
                    rec = (not r1 or t1 == r1[0]) and (not r2 or t2 == r2[0])
                    emit(wh, tc + otail, label, w, rec, hint)

    return '\n'.join(lines), stats, resolved


def build_ui(rows, slim, weapons_by_name, plugsets, html_path, n_entries):
    """Refresh the embedded DATA/ICONS/PERKS consts (and entry count) in the UI.

    PERKS maps every rollable plug hash (INCLUDING enhanced variants) back to
    the spreadsheet option name — a reverse lookup so the vault view can match
    live drops the way DIM does. This never feeds wishlist emission, which
    stays base-hash-only (invariant 5).
    """
    import json
    data, icons, perk_hashes = [], {}, {}
    for w in rows:
        base = w['Name'].split('\n')[0].strip()
        variant = w['Name'].split('\n')[1].strip() if '\n' in w['Name'] else ''
        p1s, p2s = opts(w['Perk 1']), opts(w['Perk 2'])
        if not p1s and not p2s:
            continue
        barrels, mags, origins = opts(w['Barrel']), opts(w['Mag']), opts(w['Origin Trait'])
        cands = []
        for key in [norm(base)] + [norm(base) + s for s in ADEPT_SUFFIXES]:
            cands.extend(weapons_by_name.get(key, []))
        kept, wicon = [], ''
        u = {'b': set(), 'm': set(), 'p1': set(), 'p2': set()}
        for wh in cands:
            idx = _plug_name_index(slim, plugsets, wh)
            canon = {n: sorted(lst)[0][1] for n, lst in idx.items()}
            if (p1s and not any(norm(o) in canon for o in p1s)):
                continue
            if (p2s and not any(norm(o) in canon for o in p2s)):
                continue
            kept.append(wh)
            wicon = wicon or slim[wh]['icon']
            for srcl, key in [(barrels, 'b'), (mags, 'm'), (p1s, 'p1'), (p2s, 'p2')]:
                for o in srcl:
                    if norm(o) in canon:
                        u[key].add(o)
                        icons.setdefault(o, slim[canon[norm(o)]]['icon'])
                        for _enh, ph in idx[norm(o)]:
                            perk_hashes[ph] = o
            for o in origins:
                if norm(o) in canon:
                    icons.setdefault(o, slim[canon[norm(o)]]['icon'])
        if not kept:
            continue
        data.append({'s': w['sheet'], 'n': base, 'v': variant,
                     't': (w['Tier'] or '?').strip(), 'f': w['Frame'],
                     'src': w.get('Source', ''),
                     'b': [o for o in barrels if o in u['b']],
                     'm': [o for o in mags if o in u['m']],
                     'p1': [o for o in p1s if o in u['p1']],
                     'p2': [o for o in p2s if o in u['p2']],
                     'o': origins[0] if origins else '', 'h': len(kept),
                     'hs': kept, 'ic': wicon})
    html = open(html_path).read()
    html = re.sub(r'const DATA = \[.*?\];\n',
                  lambda m: 'const DATA = ' + json.dumps(data, separators=(',', ':')) + ';\n',
                  html, count=1, flags=re.S)
    html = re.sub(r'const ICONS = \{.*?\};\n',
                  lambda m: 'const ICONS = ' + json.dumps(icons, separators=(',', ':')) + ';\n',
                  html, count=1, flags=re.S)
    html = re.sub(r'const PERKS = \{.*?\};\n',
                  lambda m: 'const PERKS = ' + json.dumps(perk_hashes, separators=(',', ':')) + ';\n',
                  html, count=1, flags=re.S)
    html = re.sub(r'[\d,]+ entries', f'{n_entries:,} entries', html, count=1)
    open(html_path, 'w').write(html)
    print(f'Updated UI data in {html_path}: {len(data)} weapons, {len(icons)} perk icons, '
          f'{len(perk_hashes)} perk hashes', file=sys.stderr)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('spreadsheet', help='Path to the Endgame Analysis .xlsx')
    ap.add_argument('-o', '--output', default='EndgameAnalysis-wishlist-complete.txt')
    ap.add_argument('--ui', default=None, metavar='HTML',
                    help='Also refresh the embedded data in the ladder-UI HTML file')
    ap.add_argument('--cache-dir', default=os.path.join(tempfile.gettempdir(),
                                                        'd2-manifest-cache'),
                    help='Where to cache manifest downloads (~210 MB)')
    args = ap.parse_args()

    api_key = os.environ.get('BUNGIE_API_KEY')
    if not api_key:
        sys.exit('Set BUNGIE_API_KEY in your environment first.')

    os.makedirs(args.cache_dir, exist_ok=True)
    items, plugsets_raw = fetch_manifest(api_key, args.cache_dir)
    slim, weapons_by_name, plugsets = build_indexes(items, plugsets_raw)
    rows = read_spreadsheet(args.spreadsheet)
    print(f'Spreadsheet rows: {len(rows)}', file=sys.stderr)

    out, stats, _resolved = generate(rows, slim, weapons_by_name, plugsets)
    with open(args.output, 'w') as f:
        f.write(out)
    n_entries = sum(1 for l in out.split('\n') if l.startswith('dimwishlist:'))
    if args.ui:
        build_ui(rows, slim, weapons_by_name, plugsets, args.ui, n_entries)
    combo_total = sum(v for k, v in stats.items() if k.startswith('combo'))
    print(f"Wrote {args.output}: {stats['perfect_roll']} perfect / "
          f"{stats['full_set']} full-set / {combo_total} combo-depth / "
          f"{stats['god_roll']} god / "
          f"{stats['good_roll'] + stats['god_roll_p_good_roll']} good entries "
          f"({stats['unmatched_weapons']} unmatched weapons)", file=sys.stderr)
    for k in sorted(stats):
        if k.startswith('combo'):
            print(f'  {k}: {stats[k]}', file=sys.stderr)


if __name__ == '__main__':
    main()
